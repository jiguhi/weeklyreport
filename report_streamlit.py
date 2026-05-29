import sys
import os
import json
import time
import hmac
import base64
import hashlib
import traceback
import contextlib
from datetime import datetime, timedelta
from typing import Tuple, Dict
from io import StringIO

import numpy as np
import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook

# ==========================================
# 1. Premium Visual Styles & Layout Setup
# ==========================================
st.set_page_config(
    page_title="Naver Search AD Automated Report",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Inject modern Outfit font, sleek glowing headers, glassmorphism cards and terminal log styles
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700&family=Fira+Code:wght@400;500&display=swap');

    /* Body background and base font */
    .stApp {
        background: linear-gradient(135deg, #0e1117 0%, #161a24 50%, #1e112a 100%);
        color: #e2e8f0;
        font-family: 'Outfit', -apple-system, BlinkMacSystemFont, sans-serif;
    }

    /* Main Container Card */
    .main-header-card {
        background: rgba(255, 255, 255, 0.03);
        backdrop-filter: blur(12px);
        -webkit-backdrop-filter: blur(12px);
        border: 1px solid rgba(255, 255, 255, 0.08);
        border-radius: 20px;
        padding: 2.5rem;
        margin-bottom: 2rem;
        box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.25);
        text-align: center;
    }

    .main-title {
        font-size: 2.8rem;
        font-weight: 700;
        letter-spacing: -0.02em;
        background: linear-gradient(90deg, #ff7e5f 0%, #feb47b 50%, #ff7e5f 100%);
        background-size: 200% auto;
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        animation: shine 5s linear infinite;
        margin-bottom: 0.5rem;
    }

    .subtitle {
        color: #94a3b8;
        font-size: 1.1rem;
        font-weight: 300;
        margin-bottom: 0;
    }

    @keyframes shine {
        to { background-position: 200% center; }
    }

    /* Glassmorphism Section Cards */
    .report-card {
        background: rgba(255, 255, 255, 0.03);
        backdrop-filter: blur(16px);
        -webkit-backdrop-filter: blur(16px);
        border: 1px solid rgba(255, 255, 255, 0.08);
        border-radius: 16px;
        padding: 1.5rem;
        margin-bottom: 1.5rem;
        box-shadow: 0 8px 24px 0 rgba(0, 0, 0, 0.15);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
    }

    .report-card:hover {
        transform: translateY(-2px);
        background: rgba(255, 255, 255, 0.05);
        border-color: rgba(255, 255, 255, 0.15);
        box-shadow: 0 12px 30px 0 rgba(0, 0, 0, 0.3);
    }

    .card-title {
        font-size: 1.25rem;
        font-weight: 600;
        color: #f1f5f9;
        border-left: 4px solid #ff7e5f;
        padding-left: 10px;
        margin-bottom: 1.2rem;
    }

    /* Terminal Console Box */
    .console-box {
        background: #0a0b10;
        border: 2px solid #f59e0b;
        border-radius: 12px;
        box-shadow: 0 0 20px rgba(245, 158, 11, 0.15);
        overflow: hidden;
        margin-top: 1.5rem;
        font-family: 'Fira Code', 'JetBrains Mono', Consolas, monospace;
    }

    .console-header {
        background: #171821;
        padding: 10px 16px;
        display: flex;
        align-items: center;
        border-bottom: 1px solid #232433;
    }

    .console-title {
        color: #34d399;
        font-size: 0.85rem;
        margin-left: 12px;
        font-weight: 600;
        letter-spacing: 0.05em;
    }

    .dot {
        width: 10px;
        height: 10px;
        border-radius: 50%;
        margin-right: 6px;
        display: inline-block;
    }

    .dot.red { background: #ef4444; }
    .dot.yellow { background: #f59e0b; }
    .dot.green { background: #10b981; }

    .console-content {
        padding: 16px;
        font-size: 0.85rem;
        color: #fef08a;
        max-height: 300px;
        overflow-y: auto;
        white-space: pre-wrap;
        line-height: 1.6;
    }

    /* Custom scrollbar for console */
    .console-content::-webkit-scrollbar {
        width: 8px;
    }
    .console-content::-webkit-scrollbar-track {
        background: #0a0b10;
    }
    .console-content::-webkit-scrollbar-thumb {
        background: #232433;
        border-radius: 4px;
    }
    .console-content::-webkit-scrollbar-thumb:hover {
        background: #f59e0b;
    }

    label,
    .stTextInput label,
    .stTextInput label p,
    .stDateInput label,
    .stDateInput label p,
    .stSelectbox label,
    .stSelectbox label p,
    .stTextArea label,
    .stTextArea label p {
        color: #111827 !important;
        -webkit-text-fill-color: #111827 !important;
    }

    [data-testid="stMain"] .stTextInput label,
    [data-testid="stMain"] .stTextInput label p,
    [data-testid="stMain"] .stTextArea label,
    [data-testid="stMain"] .stTextArea label p {
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
    }

    .stTextInput input,
    .stDateInput input,
    .stTextArea textarea,
    [data-testid="stSidebar"] input,
    [data-testid="stSidebar"] textarea {
        color: #111827 !important;
        -webkit-text-fill-color: #111827 !important;
    }

    .stTextInput input::placeholder,
    .stDateInput input::placeholder,
    .stTextArea textarea::placeholder,
    [data-testid="stSidebar"] input::placeholder,
    [data-testid="stSidebar"] textarea::placeholder {
        color: rgba(17, 24, 39, 0.55) !important;
        -webkit-text-fill-color: rgba(17, 24, 39, 0.55) !important;
    }

    .stButton button,
    .stButton button p,
    .stDownloadButton button,
    .stDownloadButton button p {
        color: #111827 !important;
        -webkit-text-fill-color: #111827 !important;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# ==========================================
# 2. Local Configuration Persistence
# ==========================================
CONFIG_PATH = os.path.join(os.path.dirname(__file__), ".report_config.json")

def load_config() -> dict:
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_config(config_dict: dict):
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(config_dict, f, ensure_ascii=False, indent=4)
    except Exception:
        pass

# Initialize session state using loaded config
config_data = load_config()

# Default values mapping
defaults = {
    "customer_id": "1596292",
    "api_key": "",
    "secret_key": "",
    "r1_start": (datetime.now() - timedelta(days=7)).strftime("%Y%m%d"),
    "r1_end": (datetime.now() - timedelta(days=1)).strftime("%Y%m%d"),
    "r1_excel": "c:/Users/User/Desktop/code/report.xlsx",
    "r1_sheet": "일별보고서",
    "r2_start": (datetime.now() - timedelta(days=7)).strftime("%Y%m%d"),
    "r2_end": (datetime.now() - timedelta(days=1)).strftime("%Y%m%d"),
    "r2_excel": "c:/Users/User/Desktop/code/report.xlsx",
    "r2_sheet": "소재별보고서",
    "r3_start": (datetime.now() - timedelta(days=7)).strftime("%Y%m%d"),
    "r3_end": (datetime.now() - timedelta(days=1)).strftime("%Y%m%d"),
    "r3_excel": "c:/Users/User/Desktop/code/report.xlsx",
    "r3_sheet": "검색어보고서",
    "all_start": (datetime.now() - timedelta(days=7)).strftime("%Y%m%d"),
    "all_end": (datetime.now() - timedelta(days=1)).strftime("%Y%m%d"),
    "all_excel": "c:/Users/User/Desktop/code/report.xlsx"
}

for key, default_val in defaults.items():
    if key not in st.session_state:
        st.session_state[key] = config_data.get(key, default_val)

# Callback to persist state dynamically
def persist_state():
    current_config = {}
    for key in defaults.keys():
        if key in st.session_state:
            current_config[key] = st.session_state[key]
    save_config(current_config)

# ==========================================
# 3. In-memory Realtime Logging Console
# ==========================================
class StreamlitLogBuffer:
    def __init__(self, placeholder):
        self.placeholder = placeholder
        self.buffer = []
        self.max_lines = 100

    def write(self, text):
        if text:
            # Output to actual console first
            sys.__stdout__.write(text)
            self.buffer.append(text)
            if len(self.buffer) > self.max_lines:
                self.buffer.pop(0)
            
            # Format and display in Streamlit
            logs_html = "".join(self.buffer).replace("\n", "<br>")
            self.placeholder.markdown(
                f"""
                <div class="console-box">
                    <div class="console-header">
                        <span class="dot red"></span>
                        <span class="dot yellow"></span>
                        <span class="dot green"></span>
                        <span class="console-title">Console Logs</span>
                    </div>
                    <div class="console-content">
                        {logs_html}
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )

    def flush(self):
        sys.__stdout__.flush()

@contextlib.contextmanager
def capture_stdout(placeholder):
    buffer = StreamlitLogBuffer(placeholder)
    old_stdout = sys.stdout
    old_stderr = sys.stderr
    sys.stdout = buffer
    sys.stderr = buffer
    try:
        yield buffer
    finally:
        sys.stdout = old_stdout
        sys.stderr = old_stderr

# ==========================================
# 4. Signature & Report Generation Core Logic
# ==========================================
class Signature:
    @staticmethod
    def generate(timestamp, method, uri, secret_key):
        message = f"{timestamp}.{method}.{uri}"
        hash_ = hmac.new(bytes(secret_key, "utf-8"), bytes(message, "utf-8"), hashlib.sha256)
        return base64.b64encode(hash_.digest()).decode()

    @staticmethod
    def get_header(method, uri, api_key, secret_key, customer_id):
        timestamp = str(round(time.time() * 1000))
        signature = Signature.generate(timestamp, method, uri, secret_key)
        return {
            "Content-Type": "application/json; charset=UTF-8",
            "X-Timestamp": timestamp,
            "X-API-KEY": api_key,
            "X-Customer": str(customer_id),
            "X-Signature": signature,
        }

class NaverReportRunner:
    def __init__(self, api_key, secret_key, customer_id):
        self.API_KEY = api_key.strip()
        self.SECRET_KEY = secret_key.strip()
        self.CUSTOMER_ID = customer_id.strip()

        self.media_code_map = {
            684924 : "네이버 통합검색 네이버플러스 스토어 - 모바일",
            684925 : "네이버 통합검색 네이버플러스 스토어 - PC",
            684926 : "네이버플러스 스토어 - 모바일",
            684927 : "네이버플러스 스토어 - PC",
            783980 : "네이버플러스 스토어 검색창 - 모바일",
            783981 : "네이버플러스 스토어 검색창 - PC",
            783978 : "네이버 쇼핑 검색창 - 모바일",
            667909 : "네이버 메인 - 모바일 홈피드",
            475844 : "네이버 쇼핑 탭 추천 - 모바일",
            623353 : "네이버 쇼핑 스마트스토어 추천 - 모바일",
            643599 : "네이버 쇼핑 스마트스토어 추천 - PC",
            341893 : "네이버 통합검색 추천 - 모바일",
            341898 : "네이버 쇼핑 추천 - 모바일",
            370822 : "네이버 통합검색 추천 - PC",
            556927 : "네이버 쇼핑 추천 - PC",
            370824 : "네이버 쇼핑 카탈로그 추천 - 모바일",
            370826 : "네이버 쇼핑 카탈로그 추천 - PC",
            783979 : "네이버 쇼핑 검색창 - PC",
            27758: "네이버 통합검색 - PC",
            8753: "네이버 통합검색 - 모바일",
            122876: "네이버 검색탭",
            122875: "네이버 통합검색 광고더보기",
            11068: "네이버 쇼핑 - PC",
            33421: "네이버 쇼핑 - 모바일",
            1525: "네이버 지식iN - PC",
            36010: "네이버 지식iN - 모바일",
            96499: "네이버 카페 - PC",
            96500: "네이버 카페 - 모바일",
            118495: "ZUM - PC",
            118496: "ZUM - 모바일",
            700903: "네이버 통합검색 추천형 피드지면 - 모바일",
            103848: "밴드(BAND) - 모바일",
            38367: "11번가 - PC",
            38630: "11번가 - 모바일",
            37853: "2CPU",
            23650: "82cook",
            335739: "Bing - 모바일",
            335738: "Bing - PC",
            769945: "DVD프라임 - 모바일",
            245885: "MLB파크 - PC",
            245891: "MLB파크 - 모바일",
            419964: "OK캐쉬백 - 모바일",
            593822: "SLR클럽 - PC",
            593823: "SLR클럽 - 모바일",
            66998: "YTN - PC",
            67582: "YTN - 모바일",
            81750: "가생이닷컴",
            15121: "간호잡",
            718848: "개드립넷 - PC",
            718849: "개드립넷 - 모바일",
            58824: "건설워커 - PC",
            74321: "건설워커 - 모바일",
            49749: "교차로 - 모바일",
            502366: "궁금하넷 - 모바일",
            128029: "꼬망세",
            190272: "나무위키 - PC",
            190273: "나무위키 - 모바일",
            731553: "네이트 - PC",
            731554: "네이트 - 모바일",
            609156: "노써치 - 모바일",
            774514: "뉴닉 - PC",
            655425: "뉴덕 - PC",
            429046: "뉴스펍 - 모바일",
            298463: "뉴스픽 - 모바일",
            23123: "다나와 - PC",
            87620: "다나와 - 모바일",
            612593: "다음 - PC",
            612594: "다음 - 모바일",
            742014: "다음카페 - PC",
            742015: "다음카페 - 모바일",
            168665: "다이닝코드 - PC",
            168666: "다이닝코드 - 모바일",
            769266: "닥터나우 - 모바일",
            593821: "더팀스 - 모바일",
            605402: "디미토리 - PC",
            605403: "디미토리 - 모바일",
            718850: "디시이슈 - 모바일",
            139215: "디시인사이드 - PC",
            131019: "디시인사이드 - 모바일",
            707083: "디시트렌드 - 모바일",
            67000: "디자이너잡",
            788865: "라이너 - PC",
            141121: "레포트샵",
            41352: "레포트월드",
            643228: "로톡 - 모바일",
            151173: "루리웹 - PC",
            151174: "루리웹 - 모바일",
            51655: "마이민트",
            137282: "마이민트 - 모바일",
            35324: "마이클럽",
            26506: "맘스다이어리",
            707082: "매니아 - 모바일",
            58827: "메디업 - PC",
            62767: "메디업 - 모바일",
            37126: "메디잡 - PC",
            74320: "메디잡 - 모바일",
            58825: "메디컬잡",
            56345: "미디어잡",
            98128: "번개장터 - 모바일",
            248824: "번개장터 - PC",
            15124: "벼룩시장 - PC",
            54186: "벼룩시장 - 모바일",
            282742: "보배드림 - 모바일",
            16334: "부동산써브",
            84644: "비즈폼",
            27567: "뽐뿌 - PC",
            49745: "뽐뿌 - 모바일",
            69559: "사람인 - 모바일",
            185344: "사람인 - PC",
            69555: "샵마넷 - PC",
            69561: "샵마넷 - 모바일",
            69557: "샵오픈",
            779437: "세리에매니아 - 모바일",
            156872: "셀잇 - PC",
            156873: "셀잇 - 모바일",
            242241: "속닥 - 모바일",
            141763: "쇼킹딜 - 모바일",
            62766: "수다닷컴",
            386648: "순위닷 - 모바일",
            417067: "시럽월렛 - 모바일",
            151175: "씽크존 - PC",
            20545: "아이베이비 - PC",
            398851: "안랩V3 - PC",
            24087: "알바몬",
            238734: "알바몬 - 모바일",
            15119: "알바천국 - PC",
            49746: "알바천국 - 모바일",
            36379: "에누리닷컴 - PC",
            45714: "에누리닷컴 - 모바일",
            479286: "에브리타임 - 모바일",
            137280: "에펨코리아 - PC",
            137281: "에펨코리아 - 모바일",
            79387: "여행오키",
            502365: "열달후에 - 모바일",
            38193: "예스폼",
            70389: "오늘의유머",
            443152: "오락 - 모바일",
            1526: "옥션 - PC",
            131018: "옥션 - 모바일",
            131268: "옥션중고장터 - 모바일",
            162341: "와글바글",
            689351: "웃긴대학 - PC",
            49363: "웃긴대학 - 모바일",
            58826: "이엔지잡",
            37131: "이지데이 - PC",
            49747: "이지데이 - 모바일",
            417064: "이토랜드 - PC",
            417065: "이토랜드 - 모바일",
            484086: "인스티즈 - 모바일",
            38197: "인크루트 - PC",
            56346: "인크루트 - 모바일",
            16333: "인터넷교차로",
            38628: "일간스포츠",
            28552: "잡코리아 - PC",
            51271: "잡코리아 - 모바일",
            291244: "중고나라 - 모바일",
            15604: "지식로그",
            291245: "최저가마켓 - 모바일",
            645386: "침하하 - 모바일",
            666914: "캐시닥 - 모바일",
            756085: "캐시슬라이드 - 모바일",
            727930: "캐시워크 - 모바일",
            655426: "캐시워크뉴스 - 모바일",
            282741: "캐시피드 - 모바일",
            762004: "케어파트너 - 모바일",
            400341: "코리아닷컴 - PC",
            598347: "콴다 - 모바일",
            20546: "쿠차 - PC",
            172112: "쿠차 - 모바일",
            39237: "쿠폰모아 - PC",
            190274: "쿠폰모아 - 모바일",
            19369: "클리앙 - PC",
            137283: "클리앙 - 모바일",
            15122: "키드키즈",
            785950: "키즈노트 - 모바일",
            242240: "키즈맘 - PC",
            242239: "텐아시아 - PC",
            484085: "티맵 - 모바일",
            735234: "티스토리 - PC",
            735235: "티스토리 - 모바일",
            69558: "패션워크",
            429047: "풀빵닷컴 - 모바일",
            762003: "플래텀 - PC",
            484542: "한겨레신문 - 모바일",
            20049: "한경닷컴 - PC",
            51591: "한경닷컴 - 모바일",
            106391: "해피캠퍼스 - PC",
            106392: "해피캠퍼스 - 모바일",
            156874: "해피학술 - PC",
            417070: "행복쇼핑 - PC",
            417071: "행복쇼핑 - 모바일",
            498105: "헬로마켓 - 모바일",
            49362: "호텔모아",
            791888: "에누리 - 모바일",
            791889: "에누리 - PC",
            805760: "네이트검색-모바일",
            684930: "네이버플러스 스토어베스트 - 모바일",
            684931: "네이버플러스 스토어베스트 - PC",
            805759: "네이트검색-PC", 
            787480: "네이버 선물 검색 - PC",
            787478: "네이버 통합검색 선물 - PC",
            787479: "네이버 선물 검색 - 모바일",
            787481: "네이버 통합검색 선물 - 모바일",
            424040: "네이버 검색탭 - 모바일",

        }

        self.Adgroup_Type_map = {
            "WEB_SITE": "파워링크",
            "SHOPPING": "쇼핑검색",
            "INFORMATION": "파워컨텐츠(정보)",
            "PRODUCT": "파워컨텐츠(상품)",
            "BRAND_SEARCH": "브랜드검색",
            "PLACE": "플레이스",
            "CATALOG": "카탈로그",
            "SHOPPING_BRAND": "브랜드쇼핑",
            "LOCAL_AD": "지역소상공인",
            "BRAND_NEW": "신제품검색광고"
        }

    def log(self, msg: str):
        print(msg)

    def daterange_yyyymmdd(self, start_yyyymmdd: str, end_yyyymmdd: str):
        s = datetime.strptime(start_yyyymmdd, "%Y%m%d")
        e = datetime.strptime(end_yyyymmdd, "%Y%m%d")
        cur = s
        while cur <= e:
            yield cur.strftime("%Y%m%d")
            cur += timedelta(days=1)

    def create_and_download_report_statdt(self, reportTp: str, statDt: str) -> str:
        BASE_URL = "https://api.searchad.naver.com"
        uri = "/stat-reports"
        method = "POST"
        payload = {"reportTp": reportTp, "statDt": statDt}

        r = requests.post(
            BASE_URL + uri,
            json=payload,
            headers=Signature.get_header(method, uri, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID),
            timeout=30,
        )
        r.raise_for_status()

        job = r.json()
        reportJobId = job.get("reportJobId")
        if not reportJobId:
            raise RuntimeError(f"reportJobId 없음: {job}")

        uri_get = f"/stat-reports/{reportJobId}"
        method_get = "GET"
        downloadUrl = ""
        last_info = None

        for _ in range(90):
            g = requests.get(
                BASE_URL + uri_get,
                headers=Signature.get_header(method_get, uri_get, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID),
                timeout=30,
            )
            g.raise_for_status()
            info = g.json()
            last_info = info
            status = (info.get("status") or "").upper()
            downloadUrl = (info.get("downloadUrl") or "").strip()

            if status in ["NONE", "FAIL", "ERROR"]:
                self.log(f"[{reportTp}] statDt={statDt} status={status} -> 빈 리포트 처리 (job={reportJobId})")
                return ""

            if downloadUrl:
                break
            time.sleep(1)

        if not downloadUrl:
            self.log(f"[{reportTp}] downloadUrl 생성 실패 -> 빈 리포트 처리 (job={reportJobId}, last={last_info})")
            return ""

        from urllib.parse import urlsplit, parse_qsl
        qs = dict(parse_qsl(urlsplit(downloadUrl).query, keep_blank_values=True))
        token = qs.get("authtoken")
        file_version = qs.get("fileVersion") or qs.get("fileversion") or "v2"

        if not token:
            self.log(f"[{reportTp}] downloadUrl에 authtoken 없음 -> 빈 리포트 처리 (job={reportJobId}, url={downloadUrl})")
            return ""

        uri_dl = "/report-download"
        method_dl = "GET"
        dl = requests.get(
            BASE_URL + uri_dl,
            params={"authtoken": token, "fileVersion": file_version},
            headers=Signature.get_header(method_dl, uri_dl, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID),
            timeout=60,
        )

        if dl.status_code >= 400:
            self.log(f"[{reportTp}] report-download 실패({dl.status_code}) -> 빈 리포트 처리 (job={reportJobId})")
            return ""

        return dl.text or ""

    def create_and_download_report_range(self, reportTp: str, startDt: str, endDt: str) -> str:
        BASE_URL = 'https://api.searchad.naver.com'
        uri = "/stat-reports"
        method = "POST"
        payload = {"reportTp": reportTp, "startDt": startDt, "endDt": endDt}

        r = requests.post(
            BASE_URL + uri, 
            json=payload,
            headers=Signature.get_header(method, uri, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID)
        )
        r.raise_for_status()

        job = r.json()
        reportJobId = job.get("reportJobId")
        if not reportJobId:
            raise RuntimeError(f"reportJobId 없음: {job}")

        uri_get = f"/stat-reports/{reportJobId}"
        method_get = "GET"
        downloadUrl = ""
        for _ in range(90):
            g = requests.get(
                BASE_URL + uri_get,
                headers=Signature.get_header(method_get, uri_get, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID)
            )
            g.raise_for_status()
            info = g.json()
            downloadUrl = info.get("downloadUrl") or ""
            if downloadUrl:
                break
            time.sleep(1)

        if not downloadUrl:
            raise RuntimeError(f"downloadUrl 생성 실패. reportJobId={reportJobId}")

        from urllib.parse import urlsplit, parse_qsl
        qs = dict(parse_qsl(urlsplit(downloadUrl).query, keep_blank_values=True))
        token = qs.get("authtoken")
        file_version = qs.get("fileVersion") or qs.get("fileversion") or "v2"

        uri_dl = "/report-download"
        method_dl = "GET"
        dl = requests.get(
            BASE_URL + uri_dl,
            params={"authtoken": token, "fileVersion": file_version},
            headers=Signature.get_header(method_dl, uri_dl, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID)
        )
        dl.raise_for_status()
        return dl.text

    def _append_df_to_excel(self, df: pd.DataFrame, excel_path: str, sheet_name: str):
        os.makedirs(os.path.dirname(excel_path) or ".", exist_ok=True)

        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                startrow = ws.max_row
                header = False
            else:
                wb.create_sheet(sheet_name)
                wb.save(excel_path)
                startrow = 0
                header = True

            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    header=header,
                    startrow=startrow
                )
        else:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    def run_daily_report(self, start_date, end_date, sheet_name, excel):
        self.log("일별 보고서생성 시작!")
        BASE_URL = 'https://api.searchad.naver.com'

        # 1) Campaigns Mapping
        uri = "/ncc/campaigns"
        resp = requests.get(BASE_URL + uri, headers=Signature.get_header("GET", uri, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID))
        resp.raise_for_status()
        campaigns = resp.json()
        df_campaign = pd.DataFrame(campaigns)[["nccCampaignId", "name"]].rename(
            columns={"nccCampaignId": "Campaign ID", "name": "Campaign Name"}
        )

        # 2) Adgroups Mapping
        uri_ag = "/ncc/adgroups"
        adgroups_all = []
        for cid in df_campaign["Campaign ID"].tolist():
            r_ag = requests.get(
                BASE_URL + uri_ag,
                params={"nccCampaignId": cid},
                headers=Signature.get_header("GET", uri_ag, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID)
            )
            if r_ag.status_code == 200:
                adgroups_all.extend(r_ag.json())

        df_ag = pd.DataFrame(adgroups_all)[["nccAdgroupId", "name", "nccCampaignId", "adgroupType"]].rename(
            columns={
                "nccAdgroupId": "AD Group ID",
                "name": "AD Group Name",
                "nccCampaignId": "Campaign ID",
                "adgroupType": "Adgroup Type"
            }
        )

        # 3) Ads Mapping
        uri_ad = "/ncc/ads"
        ads_all = []
        for agid in df_ag["AD Group ID"].dropna().unique().tolist():
            r_ad = requests.get(
                BASE_URL + uri_ad,
                params={"nccAdgroupId": agid},
                headers=Signature.get_header("GET", uri_ad, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID)
            )
            if r_ad.status_code == 200:
                ads_all.extend(r_ad.json())

        df_ad = pd.DataFrame(ads_all)[["nccAdId", "nccAdgroupId"]].rename(
            columns={"nccAdId": "AD ID", "nccAdgroupId": "AD Group ID"}
        )

        df_cam = (
            df_ag.merge(df_campaign[["Campaign ID", "Campaign Name"]], on="Campaign ID", how="left")
                .merge(df_ad, on="AD Group ID", how="left")
                [["Campaign Name","Campaign ID","AD Group Name","AD Group ID","AD ID","Adgroup Type"]]
        )

        # 4) AD report downloads
        column_names2 = [
            "Date", "Customer_ID", "Campaign ID","AD Group ID","AD keyword ID","AD ID","Business Channel ID",
            "Media Code","PC Mobile Type","Impression","Click","Cost","Sum of Ad Rank","View Count"
        ]

        frames = []
        for d in self.daterange_yyyymmdd(start_date, end_date):
            ad_text = self.create_and_download_report_statdt("AD", d)
            df_day = pd.read_csv(StringIO(ad_text), sep="\t", header=None, names=column_names2)
            frames.append(df_day)

        df_report2 = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=column_names2)

        df_report2["Cost"] = pd.to_numeric(df_report2["Cost"], errors="coerce").fillna(0) 
        df_report2["Impression"] = pd.to_numeric(df_report2["Impression"], errors="coerce").fillna(0)
        df_report2["Click"] = pd.to_numeric(df_report2["Click"], errors="coerce").fillna(0)
        df_report2['Average of Rank'] = df_report2['Sum of Ad Rank'] / df_report2['Impression']
        df_report2 = df_report2.round({'Average of Rank' : 1})

        df_report2 = df_report2.merge(
            df_cam.drop_duplicates(subset=["Campaign ID","AD Group ID","AD ID"]),
            on=["Campaign ID","AD Group ID","AD ID"],
            how="left"
        )

        for c in ["Campaign Name","AD Group Name","Adgroup Type"]:
            df_report2[c] = df_report2[c].fillna("UNKNOWN")

        df_report_group2 = df_report2.groupby(
                ["Date","Campaign Name","Campaign ID","AD Group ID","AD Group Name","Adgroup Type","AD ID","Media Code","PC Mobile Type"],
                as_index=False
            ) .agg({
                "Impression": "sum",
                "Click": "sum",
                "Cost": "sum",
                "Sum of Ad Rank": "sum",
                "View Count": "sum",
                "Average of Rank": "mean",
            })

        # 5) Conversion reports
        column_names = [
            "Date","CUSTOMER ID","Campaign ID","AD Group ID","AD keyword ID","AD ID","Business Channel ID",
            "Hours","Region code","Media code","PC Mobile Type",
            "Conversion Method","Conversion Type","Conversion count","Sales by conversion"
        ]

        df_report = pd.DataFrame(columns=column_names)
        try:
            conv_text = self.create_and_download_report_range("AD_CONVERSION_DETAIL", start_date, end_date)
            df_report = pd.read_csv(StringIO(conv_text), sep="\t", header=None, names=column_names)
        except Exception as e:
            self.log(f"[WARN] range 전환리포트 실패({start_date}~{end_date}) → statDt로 fallback. err={e}")

            frames = []
            for d in self.daterange_yyyymmdd(start_date, end_date):
                try:
                    conv_text = self.create_and_download_report_statdt("AD_CONVERSION_DETAIL", d)
                    df_day = pd.read_csv(StringIO(conv_text), sep="\t", header=None, names=column_names)
                    frames.append(df_day)
                    self.log(f'전환리포트(statDt) 성공 : {d}')
                except Exception as e2:
                    self.log(f"[SKIP] 전환리포트(statDt) 실패: {d}, err={e2}")
                    continue

            if frames:
                df_report = pd.concat(frames, ignore_index=True)

        df_report = df_report.merge(
            df_cam.drop_duplicates(subset=["Campaign ID","AD Group ID","AD ID"]),
            on=["Campaign ID","AD Group ID","AD ID"],
            how="left"
        )

        for c in ["Campaign Name","AD Group Name","Adgroup Type"]:
            df_report[c] = df_report[c].fillna("UNKNOWN")

        df_report_group = self._group_conversion_breakdown(
            df_report,
            ["Date","Campaign Name","Campaign ID","AD Group ID","AD Group Name","Adgroup Type","AD ID","Media code","PC Mobile Type"]
        )

        df_report_group = df_report_group.rename(columns={"Media code":"Media Code"})

        # 6) Merge
        merge_keys = ["Date","Campaign Name","Campaign ID","AD Group ID","AD Group Name","Adgroup Type","AD ID","Media Code","PC Mobile Type"]

        for k in merge_keys:
            df_report_group[k]  = df_report_group[k].astype(str).str.strip()
            df_report_group2[k] = df_report_group2[k].astype(str).str.strip()

        df_combined = pd.merge(df_report_group, df_report_group2, on=merge_keys, how="outer")

        for c in ["Total conversion count","Total sales by conversion","Purchase conversion count",
                  "Purchase sales by conversion","Cart conversion count","Cart sales by conversion"]:
            df_combined[c] = pd.to_numeric(df_combined[c], errors="coerce").fillna(0)
        df_combined["Impression"] = pd.to_numeric(df_combined["Impression"], errors="coerce").fillna(0)
        df_combined["Click"] = pd.to_numeric(df_combined["Click"], errors="coerce").fillna(0)
        df_combined["Cost"] = pd.to_numeric(df_combined["Cost"], errors="coerce").fillna(0)
        df_combined["Average of Rank"] = pd.to_numeric(df_combined["Average of Rank"], errors="coerce").fillna(0)
        df_combined["Media Name"] = pd.to_numeric(df_combined["Media Code"], errors="coerce").map(self.media_code_map).fillna("기타 매체")
        df_combined["Adgroup Type"] = df_combined["Adgroup Type"].astype(str).map(self.Adgroup_Type_map).fillna("기타")

        df_combined = df_combined.drop(columns=["Campaign ID","AD Group ID","Media Code"], errors="ignore")

        df_combined = df_combined.rename(columns={
            "Date": "일별",
            "Adgroup Type": "캠페인유형",
            "Campaign Name": "캠페인",
            "AD Group Name": "광고그룹",
            "AD ID": "소재",
            "PC Mobile Type": "PC/Mo",
            "Media Name": "매체이름",
            "Impression": "노출수",
            "Click": "클릭수",
            "Cost": "총비용(VAT포함,원)",
            "Total conversion count": "총 전환수",
            "Total sales by conversion": "총 전환매출액(원)",
            "Purchase conversion count": "구매완료 전환수",
            "Purchase sales by conversion": "구매완료 전환매출액(원)",
            "Cart conversion count": "장바구니 전환수",
            "Cart sales by conversion": "장바구니 전환매출액(원)",
            "Average of Rank": "평균 노출 순위"
        })
        sd = pd.to_datetime(start_date, format="%Y%m%d")

        month_label = f"{sd.month}월"
        week_of_month = ((sd.day - 1) // 7) + 1
        해당주차 = f"{week_of_month}주차"

        df_final = (
            df_combined
            .groupby([
                "일별",
                "캠페인유형",
                "캠페인",
                "광고그룹",
                "소재",
                "PC/Mo"
            ], as_index=False)
            .agg({
                "노출수": "sum",
                "클릭수": "sum",
                "총비용(VAT포함,원)": "sum",
                "총 전환수": "sum",
                "총 전환매출액(원)": "sum",
                "구매완료 전환수": "sum",
                "구매완료 전환매출액(원)": "sum",
                "장바구니 전환수": "sum",
                "장바구니 전환매출액(원)": "sum",
                "평균 노출 순위": lambda x: np.average(
                    x,
                    weights=df_combined.loc[x.index, "노출수"]
                ) if df_combined.loc[x.index, "노출수"].sum() > 0 else np.nan
            })
        )
        if "평균 노출 순위" in df_final.columns:
            df_final["평균 노출 순위"] = (
                pd.to_numeric(df_final["평균 노출 순위"], errors="coerce")
                .round(1)
            )
        df_final = df_final[[
            "일별","캠페인유형","캠페인","광고그룹","소재","PC/Mo",
            "노출수","클릭수","총비용(VAT포함,원)",
            "총 전환수","총 전환매출액(원)",
            "구매완료 전환수","구매완료 전환매출액(원)",
            "평균 노출 순위",
            "장바구니 전환수","장바구니 전환매출액(원)"
        ]]
        df_final = df_final.sort_values("일별")
        df_final["일별"] = pd.to_datetime(df_final["일별"], format="%Y%m%d").dt.normalize()

        df_final.insert(0, "해당월", month_label)
        df_final.insert(1, "해당주차", 해당주차)
        df_final.insert(2, "   ", "")
        df_final.insert(3, "    ", "")
        df_final.insert(4, "     ", "")

        self._append_df_to_excel(df_final, excel, sheet_name)
        self.log("일별데이터 저장완료!ㅇoㅇ")

    def run_ad_report(self, start_date, end_date, sheet_name, excel):
        self.log("소재별 보고서생성 시작!")
        BASE_URL = 'https://api.searchad.naver.com'

        # 1) Campaigns Mapping
        uri = "/ncc/campaigns"
        resp = requests.get(BASE_URL + uri, headers=Signature.get_header("GET", uri, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID))
        resp.raise_for_status()
        campaigns = resp.json()
        df_campaign = pd.DataFrame(campaigns)[["nccCampaignId", "name"]].rename(
            columns={"nccCampaignId": "Campaign ID", "name": "Campaign Name"}
        )

        # 2) Adgroups Mapping
        uri_ag = "/ncc/adgroups"
        adgroups_all = []
        for cid in df_campaign["Campaign ID"].tolist():
            r_ag = requests.get(
                BASE_URL + uri_ag,
                params={"nccCampaignId": cid},
                headers=Signature.get_header("GET", uri_ag, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID)
            )
            if r_ag.status_code == 200:
                adgroups_all.extend(r_ag.json())

        df_ag = pd.DataFrame(adgroups_all)[["nccAdgroupId", "name", "nccCampaignId", "adgroupType"]].rename(
            columns={
                "nccAdgroupId": "AD Group ID",
                "name": "AD Group Name",
                "nccCampaignId": "Campaign ID",
                "adgroupType": "Adgroup Type"
            }
        )

        # 3) Ads Mapping
        uri_ad = "/ncc/ads"
        ads_all = []
        for agid in df_ag["AD Group ID"].dropna().unique().tolist():
            r_ad = requests.get(
                BASE_URL + uri_ad,
                params={"nccAdgroupId": agid},
                headers=Signature.get_header("GET", uri_ad, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID)
            )
            if r_ad.status_code == 200:
                ads_all.extend(r_ad.json())

        df_ad = pd.DataFrame(ads_all)[["nccAdId", "nccAdgroupId"]].rename(
            columns={"nccAdId": "AD ID", "nccAdgroupId": "AD Group ID"}
        )

        df_cam = (
            df_ag.merge(df_campaign[["Campaign ID", "Campaign Name"]], on="Campaign ID", how="left")
                .merge(df_ad, on="AD Group ID", how="left")
                [["Campaign Name","Campaign ID","AD Group Name","AD Group ID","AD ID","Adgroup Type"]]
        )

        # 4) AD Report downloads
        column_names2 = [
            "Date", "Customer_ID", "Campaign ID","AD Group ID","AD keyword ID","AD ID","Business Channel ID",
            "Media Code","PC Mobile Type","Impression","Click","Cost","Sum of Ad Rank","View Count"
        ]

        frames = []
        for d in self.daterange_yyyymmdd(start_date, end_date):
            ad_text = self.create_and_download_report_statdt("AD", d)
            df_day = pd.read_csv(StringIO(ad_text), sep="\t", header=None, names=column_names2)
            frames.append(df_day)

        df_report2 = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=column_names2)

        df_report2["Cost"] = pd.to_numeric(df_report2["Cost"], errors="coerce").fillna(0) 
        df_report2["Impression"] = pd.to_numeric(df_report2["Impression"], errors="coerce").fillna(0)
        df_report2["Click"] = pd.to_numeric(df_report2["Click"], errors="coerce").fillna(0)
        df_report2['Average of Rank'] = df_report2['Sum of Ad Rank'] / df_report2['Impression']
        df_report2 = df_report2.round({'Average of Rank' : 1})

        df_report2 = df_report2.merge(
            df_cam.drop_duplicates(subset=["Campaign ID","AD Group ID","AD ID"]),
            on=["Campaign ID","AD Group ID","AD ID"],
            how="left"
        )

        for c in ["Campaign Name","AD Group Name","Adgroup Type"]:
            df_report2[c] = df_report2[c].fillna("UNKNOWN")

        df_report_group2 = df_report2.groupby(
                ["Date","Campaign Name","Campaign ID","AD Group ID","AD Group Name","Adgroup Type","AD ID","Media Code","PC Mobile Type"],
                as_index=False
            ) .agg({
                "Impression": "sum",
                "Click": "sum",
                "Cost": "sum",
                "Sum of Ad Rank": "sum",
                "View Count": "sum",
                "Average of Rank": "mean",
            })

        # 5) Conversion downloads
        column_names = [
            "Date","CUSTOMER ID","Campaign ID","AD Group ID","AD keyword ID","AD ID","Business Channel ID",
            "Hours","Region code","Media code","PC Mobile Type",
            "Conversion Method","Conversion Type","Conversion count","Sales by conversion"
        ]

        df_report = pd.DataFrame(columns=column_names)
        try:
            conv_text = self.create_and_download_report_range("AD_CONVERSION_DETAIL", start_date, end_date)
            df_report = pd.read_csv(StringIO(conv_text), sep="\t", header=None, names=column_names)
        except Exception:
            frames = []
            for d in self.daterange_yyyymmdd(start_date, end_date):
                try:
                    conv_text = self.create_and_download_report_statdt("AD_CONVERSION_DETAIL", d)
                    df_day = pd.read_csv(StringIO(conv_text), sep="\t", header=None, names=column_names)
                    frames.append(df_day)
                    self.log(f'전환리포트(statDt) 성공 : {d}')
                except Exception:
                    continue
            if frames:
                df_report = pd.concat(frames, ignore_index=True)

        df_report = df_report.merge(
            df_cam.drop_duplicates(subset=["Campaign ID","AD Group ID","AD ID"]),
            on=["Campaign ID","AD Group ID","AD ID"],
            how="left"
        )

        for c in ["Campaign Name","AD Group Name","Adgroup Type"]:
            df_report[c] = df_report[c].fillna("UNKNOWN")

        df_report_group = self._group_conversion_breakdown(
            df_report,
            ["Date","Campaign Name","Campaign ID","AD Group ID","AD Group Name","Adgroup Type","AD ID","Media code","PC Mobile Type"]
        )

        df_report_group = df_report_group.rename(columns={"Media code":"Media Code"})

        # 6) Merge
        merge_keys = ["Date","Campaign Name","Campaign ID","AD Group ID","AD Group Name","Adgroup Type","AD ID","Media Code","PC Mobile Type"]

        for k in merge_keys:
            df_report_group[k]  = df_report_group[k].astype(str).str.strip()
            df_report_group2[k] = df_report_group2[k].astype(str).str.strip()

        df_combined = pd.merge(df_report_group, df_report_group2, on=merge_keys, how="outer")

        for c in ["Total conversion count","Total sales by conversion","Purchase conversion count",
                  "Purchase sales by conversion","Cart conversion count","Cart sales by conversion"]:
            df_combined[c] = pd.to_numeric(df_combined[c], errors="coerce").fillna(0)
        df_combined["Impression"] = pd.to_numeric(df_combined["Impression"], errors="coerce").fillna(0)
        df_combined["Click"] = pd.to_numeric(df_combined["Click"], errors="coerce").fillna(0)
        df_combined["Cost"] = pd.to_numeric(df_combined["Cost"], errors="coerce").fillna(0)
        df_combined["Average of Rank"] = pd.to_numeric(df_combined["Average of Rank"], errors="coerce").fillna(0)
        df_combined["Media Name"] = pd.to_numeric(df_combined["Media Code"], errors="coerce").map(self.media_code_map).fillna("기타 매체")
        df_combined["Adgroup Type"] = df_combined["Adgroup Type"].astype(str).map(self.Adgroup_Type_map).fillna("기타")

        df_combined = df_combined.drop(columns=["Campaign ID","AD Group ID","Media Code"], errors="ignore")

        df_combined = df_combined.rename(columns={
            "Date": "일별",
            "Adgroup Type": "캠페인유형",
            "Campaign Name": "캠페인",
            "AD Group Name": "광고그룹",
            "AD ID": "소재",
            "PC Mobile Type": "PC/Mo",
            "Media Name": "매체이름",
            "Impression": "노출수",
            "Click": "클릭수",
            "Cost": "총비용(VAT포함,원)",
            "Total conversion count": "총 전환수",
            "Total sales by conversion": "총 전환매출액(원)",
            "Purchase conversion count": "구매완료 전환수",
            "Purchase sales by conversion": "구매완료 전환매출액(원)",
            "Cart conversion count": "장바구니 전환수",
            "Cart sales by conversion": "장바구니 전환매출액(원)",
            "Average of Rank": "평균 노출 순위"
        })
        value_cols = [
            "노출수","클릭수","총비용(VAT포함,원)",
            "총 전환수","총 전환매출액(원)",
            "구매완료 전환수","구매완료 전환매출액(원)",
            "장바구니 전환수","장바구니 전환매출액(원)"
        ]
        df_combined[value_cols] = df_combined[value_cols].apply(pd.to_numeric, errors="coerce").fillna(0)

        sd = pd.to_datetime(start_date, format="%Y%m%d")
        month_label = f"{sd.month}월"
        week_of_month = ((sd.day - 1) // 7) + 1
        해당주차 = f"{week_of_month}주차"

        week_result = (
            df_combined
            .drop(columns=["일별", "PC/Mo"], errors="ignore")
            .groupby(["캠페인유형","캠페인","광고그룹","소재","매체이름"], as_index=False)
            .agg({
                "노출수": "sum",
                "클릭수": "sum",
                "총비용(VAT포함,원)": "sum",
                "총 전환수": "sum",
                "총 전환매출액(원)": "sum",
                "구매완료 전환수": "sum",
                "구매완료 전환매출액(원)": "sum",
                "장바구니 전환수": "sum",
                "장바구니 전환매출액(원)": "sum",
                "평균 노출 순위" : "mean"
            })
        )
        week_result = week_result.round({'평균 노출 순위' : 1})
        week_result = week_result[[
            "캠페인유형","캠페인","광고그룹","소재","매체이름",
            "노출수","클릭수","총비용(VAT포함,원)",
            "총 전환수","총 전환매출액(원)",
            "구매완료 전환수","구매완료 전환매출액(원)",
            "평균 노출 순위",
            "장바구니 전환수","장바구니 전환매출액(원)"
        ]]

        week_result.insert(0, "해당월", month_label)
        week_result.insert(1, "해당주차", 해당주차)
        week_result.insert(2, "   ", "")
        week_result.insert(3, "    ", "")
        week_result.insert(4, "     ", "")

        self._append_df_to_excel(week_result, excel, sheet_name)
        self.log("소재별 기간 데이터 저장완료! >_<")

    def _empty_df(self, cols):
        return pd.DataFrame(columns=cols)

    def _safe_read_report_statdt(self, report_type, d, cols, sep="\t"):
        try:
            text = self.create_and_download_report_statdt(report_type, d)
            if not text or not str(text).strip():
                return self._empty_df(cols)
            return pd.read_csv(StringIO(text), sep=sep, header=None, names=cols)
        except Exception as e:
            self.log(f"[{report_type}] statDt({d}) 없음/실패 -> 빈 DF: {e}")
            return self._empty_df(cols)

    def _safe_read_report_range(self, report_type, start_date, end_date, cols, sep="\t"):
        try:
            text = self.create_and_download_report_range(report_type, start_date, end_date)
            if not text or not str(text).strip():
                return self._empty_df(cols)
            return pd.read_csv(StringIO(text), sep=sep, header=None, names=cols)
        except Exception as e:
            self.log(f"[{report_type}] range({start_date}~{end_date}) 없음/실패 -> 빈 DF: {e}")
            return self._empty_df(cols)

    def _purchase_only(self, df):
        if df.empty or "Conversion Type" not in df.columns:
            return df

        purchase_values = {"1", "1.0", "purchase", "purchasing", "purchased", "구매", "구매완료"}
        conv_type = df["Conversion Type"].astype(str).str.strip().str.lower()
        return df[conv_type.isin(purchase_values)].copy()

    def _group_purchase_conversion_like_hwanggeum(self, df, group_keys):
        result_cols = group_keys + ["Conversion count", "Sales by conversion"]
        if df.empty:
            return pd.DataFrame(columns=result_cols)

        df = df.copy()
        df["Conversion Type"] = df["Conversion Type"].fillna(0)
        df["Conversion count"] = pd.to_numeric(df["Conversion count"], errors="coerce").fillna(0)
        df["Sales by conversion"] = pd.to_numeric(df["Sales by conversion"], errors="coerce").fillna(0)

        grouped = (
            df.groupby(group_keys + ["Conversion Type"], as_index=False)
            [["Conversion count", "Sales by conversion"]].sum()
        )
        grouped["_priority"] = grouped["Conversion Type"].map({
            "purchase": 2,
            "purchasing": 2,
            "purchased": 2,
            "구매": 2,
            "구매완료": 2,
            1: 2,
            "1": 2,
            "1.0": 2,
            "add_to_cart": 1,
            0: 0,
            "0": 0,
            "0.0": 0,
        }).fillna(0)

        if grouped.empty:
            return pd.DataFrame(columns=result_cols)

        picked = grouped.loc[grouped.groupby(group_keys)["_priority"].idxmax()].reset_index(drop=True)
        picked.loc[picked["_priority"] < 2, ["Conversion count", "Sales by conversion"]] = 0

        return picked[result_cols]

    def _group_conversion_breakdown(self, df, group_keys):
        result_cols = group_keys + [
            "Total conversion count", "Total sales by conversion",
            "Purchase conversion count", "Purchase sales by conversion",
            "Cart conversion count", "Cart sales by conversion",
        ]
        if df.empty:
            return pd.DataFrame(columns=result_cols)

        df = df.copy()
        df["Conversion Type"] = df["Conversion Type"].fillna(0).astype(str).str.strip().str.lower()
        df["Conversion count"] = pd.to_numeric(df["Conversion count"], errors="coerce").fillna(0)
        df["Sales by conversion"] = pd.to_numeric(df["Sales by conversion"], errors="coerce").fillna(0)

        purchase_values = {"1", "1.0", "purchase", "purchasing", "purchased", "구매", "구매완료"}
        cart_values = {"2", "2.0", "add_to_cart", "cart", "장바구니", "장바구니담기"}

        total = (
            df.groupby(group_keys, as_index=False)[["Conversion count", "Sales by conversion"]]
            .sum()
            .rename(columns={
                "Conversion count": "Total conversion count",
                "Sales by conversion": "Total sales by conversion",
            })
        )

        purchase = (
            df[df["Conversion Type"].isin(purchase_values)]
            .groupby(group_keys, as_index=False)[["Conversion count", "Sales by conversion"]]
            .sum()
            .rename(columns={
                "Conversion count": "Purchase conversion count",
                "Sales by conversion": "Purchase sales by conversion",
            })
        )

        cart = (
            df[df["Conversion Type"].isin(cart_values)]
            .groupby(group_keys, as_index=False)[["Conversion count", "Sales by conversion"]]
            .sum()
            .rename(columns={
                "Conversion count": "Cart conversion count",
                "Sales by conversion": "Cart sales by conversion",
            })
        )

        result = total.merge(purchase, on=group_keys, how="outer").merge(cart, on=group_keys, how="outer")
        for c in result_cols:
            if c not in result.columns:
                result[c] = 0
        for c in result_cols[len(group_keys):]:
            result[c] = pd.to_numeric(result[c], errors="coerce").fillna(0)

        return result[result_cols]

    def _extract_target_value(self, target):
        if isinstance(target, str):
            text = target.strip()
            if not text:
                return ""
            try:
                target = json.loads(text)
            except Exception:
                return text

        if isinstance(target, list):
            values = [self._extract_target_value(v) for v in target]
            return ", ".join([v for v in values if v])

        if isinstance(target, dict):
            for key in ["keyword", "keywords", "searchKeyword", "searchKeywords", "requested", "value", "name"]:
                if key in target:
                    return self._extract_target_value(target.get(key))
            for value in target.values():
                extracted = self._extract_target_value(value)
                if extracted:
                    return extracted

        return ""

    def _target_field_values(self, row):
        values = []
        for key in ["targetTp", "targetType", "matchType", "keywordMatchType", "criterionType", "type"]:
            value = row.get(key)
            if value is not None:
                values.append(str(value))

        target = row.get("target") or row.get("targetJson")
        if isinstance(target, str):
            values.append(target)
            try:
                target = json.loads(target)
            except Exception:
                target = None
        if isinstance(target, dict):
            for key in ["type", "matchType", "keywordMatchType", "criterionType"]:
                value = target.get(key)
                if value is not None:
                    values.append(str(value))

        return " ".join(values).upper()

    def _normalize_match_type(self, value):
        value = str(value or "").strip()
        upper_value = value.upper()

        if any(token in upper_value for token in ["SIMILAR", "유사"]):
            return "유사일치"
        if any(token in upper_value for token in ["EXPAND", "EXTEND", "BROAD", "확장"]):
            return "확장"
        if any(token in upper_value for token in ["EXACT", "MATCH", "일치"]):
            return "일치"
        return value

    def _normalize_search_type(self, value, default=""):
        value = str(value or "").strip()
        upper_value = value.upper()

        if not value or value.lower() == "nan":
            return default
        if value in ["1", "1.0"]:
            return "일치"
        if value in ["2", "2.0"]:
            return "확장"
        if value in ["3", "3.0"]:
            return "유사일치"
        if any(token in upper_value for token in ["SIMILAR", "유사"]):
            return "유사일치"
        if any(token in upper_value for token in ["EXP", "PLUS", "BROAD", "확장"]):
            return "확장"
        if any(token in upper_value for token in ["EXACT", "MATCH", "일치"]):
            return "일치"
        return value

    def _shopping_criterion_map(self, owner_ids, base_url):
        uri = "/ncc/targets"
        rows = []

        for requested_owner_id in pd.Series(owner_ids).dropna().astype(str).str.strip().unique().tolist():
            response_rows = []
            for params in ({"ownerId": requested_owner_id}, {"ownerId": requested_owner_id, "types": "AD_TAG"}):
                try:
                    r_target = requests.get(
                        base_url + uri,
                        params=params,
                        headers=Signature.get_header("GET", uri, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID),
                        timeout=30,
                    )
                    if r_target.status_code == 200:
                        data = r_target.json()
                        response_rows = data if isinstance(data, list) else data.get("items") or data.get("data") or data.get("targets") or []
                        break
                except Exception:
                    continue

            for row in response_rows:
                owner_id = row.get("ownerId") or row.get("nccAdgroupId") or requested_owner_id
                dictionary_code = row.get("dictionaryCode") or row.get("criterionCode")
                if not owner_id or not dictionary_code:
                    continue

                target_tp = str(row.get("targetTp") or row.get("targetType") or "").strip()
                match_type = (
                    row.get("matchType")
                    or row.get("keywordMatchType")
                    or row.get("criterionType")
                    or target_tp
                )
                keyword_signal = self._target_field_values(row)
                is_ad_tag = target_tp.upper() == "AD_TAG"
                is_keyword_target = any(
                    token in keyword_signal
                    for token in ["KEYWORD", "AD_TAG", "SEARCH", "EXACT", "EXPAND", "EXTEND", "BROAD", "SIMILAR", "MATCH"]
                )
                if not (is_ad_tag or is_keyword_target):
                    continue

                target_text = self._extract_target_value(
                    row.get("target")
                    or row.get("targetJson")
                    or row.get("keyword")
                    or row.get("name")
                    or ""
                )
                normalized_match_type = "일치" if is_ad_tag else self._normalize_match_type(match_type)
                if normalized_match_type not in ["일치", "확장", "유사일치"]:
                    continue

                rows.append({
                    "Criterion id": f"{owner_id}~{dictionary_code}",
                    "Owner ID": owner_id,
                    "Search keyword": target_text or dictionary_code,
                    "Match Type": normalized_match_type,
                })

        if not rows:
            return pd.DataFrame(columns=["Criterion id", "Owner ID", "Search keyword", "Match Type"])

        return pd.DataFrame(rows).drop_duplicates(subset=["Criterion id"])

    def run_keyword_report(self, start_date, end_date, sheet_name, excel):
        self.log("키워드 보고서생성 시작!")
        BASE_URL = "https://api.searchad.naver.com"

        # 1) Campaign/Group Mapping
        uri = "/ncc/campaigns"
        resp = requests.get(BASE_URL + uri, headers=Signature.get_header("GET", uri, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID))
        resp.raise_for_status()
        campaigns = resp.json()
        df_campaign = pd.DataFrame(campaigns)[["nccCampaignId", "name"]].rename(
            columns={"nccCampaignId": "Campaign ID", "name": "Campaign Name"}
        )

        uri_ag = "/ncc/adgroups"
        adgroups_all = []
        for cid in df_campaign["Campaign ID"].tolist():
            r_ag = requests.get(
                BASE_URL + uri_ag,
                params={"nccCampaignId": cid},
                headers=Signature.get_header("GET", uri_ag, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID)
            )
            if r_ag.status_code == 200:
                adgroups_all.extend(r_ag.json())

        df_ag = pd.DataFrame(adgroups_all)[["nccAdgroupId", "name", "nccCampaignId", "adgroupType"]].rename(
            columns={
                "nccAdgroupId": "AD Group ID",
                "name": "AD Group Name",
                "nccCampaignId": "Campaign ID",
                "adgroupType": "Adgroup Type"
            }
        )

        df_cam = (
            df_ag.merge(df_campaign[["Campaign ID", "Campaign Name"]], on="Campaign ID", how="left")
                [["Campaign Name", "Campaign ID", "AD Group Name", "AD Group ID", "Adgroup Type"]]
        )

        # 2) Shopping keyword
        shopping_cols = [
            "Date", "Customer_ID", "Campaign ID","AD Group ID","Search keyword","AD ID","Business Channel ID","Hours",
            "Region code", "Media Code","PC Mobile Type","Impression","Click","Cost","Sum of Ad Rank","View Count"
        ]

        frames = []
        for d in self.daterange_yyyymmdd(start_date, end_date):
            frames.append(self._safe_read_report_statdt("SHOPPINGKEYWORD_DETAIL", d, shopping_cols))

        df_report2 = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=shopping_cols)

        df_report2["Cost"] = pd.to_numeric(df_report2["Cost"], errors="coerce").fillna(0) 
        df_report2["Impression"] = pd.to_numeric(df_report2["Impression"], errors="coerce").fillna(0)
        df_report2["Click"] = pd.to_numeric(df_report2["Click"], errors="coerce").fillna(0)
        df_report2["Sum of Ad Rank"] = pd.to_numeric(df_report2["Sum of Ad Rank"], errors="coerce").fillna(0)

        df_report2 = df_report2.merge(
            df_cam.drop_duplicates(subset=["Campaign ID","AD Group ID"]),
            on=["Campaign ID","AD Group ID"],
            how="left"
        )

        for c in ["Campaign Name","AD Group Name","Adgroup Type"]:
            df_report2[c] = df_report2[c].fillna("UNKNOWN")

        df_report_group2 = (
            df_report2.groupby(
                ["Date","Campaign Name","Campaign ID","AD Group ID","AD Group Name","Adgroup Type","Search keyword"],
                as_index=False
            )
            .agg({"Impression":"sum","Click":"sum","Cost":"sum","Sum of Ad Rank":"sum"})
        )
        df_report_group2["Average of Rank"] = np.where(
            df_report_group2["Impression"] > 0,
            df_report_group2["Sum of Ad Rank"] / df_report_group2["Impression"],
            np.nan
        ).round(1)

        conv_cols = [
            "Date","CUSTOMER ID","Campaign ID","AD Group ID","Search keyword","AD ID","Business Channel ID",
            "Hours","Region code","Media code","PC Mobile Type",
            "Conversion Method","Conversion Type","Conversion count","Sales by conversion"
        ]

        df_report = self._safe_read_report_range("SHOPPINGKEYWORD_CONVERSION_DETAIL", start_date, end_date, conv_cols)
        if df_report.empty:
            conv_frames = []
            for d in self.daterange_yyyymmdd(start_date, end_date):
                conv_frames.append(self._safe_read_report_statdt("SHOPPINGKEYWORD_CONVERSION_DETAIL", d, conv_cols))
            df_report = pd.concat(conv_frames, ignore_index=True) if conv_frames else pd.DataFrame(columns=conv_cols)

        df_report = df_report.merge(
            df_cam.drop_duplicates(subset=["Campaign ID","AD Group ID"]),
            on=["Campaign ID","AD Group ID"],
            how="left"
        )
        for c in ["Campaign Name","AD Group Name","Adgroup Type"]:
            df_report[c] = df_report[c].fillna("UNKNOWN")

        df_report_group = self._group_conversion_breakdown(
            df_report,
            ["Date","Campaign Name","Campaign ID","AD Group ID","AD Group Name","Adgroup Type","Search keyword"]
        )

        merge_keys = ["Date","Campaign Name","Campaign ID","AD Group ID","AD Group Name","Adgroup Type","Search keyword"]
        for k in merge_keys:
            df_report_group[k]  = df_report_group[k].astype(str).str.strip()
            df_report_group2[k] = df_report_group2[k].astype(str).str.strip()

        df_combined = pd.merge(df_report_group, df_report_group2, on=merge_keys, how="outer")

        for col in [
            "Total conversion count","Total sales by conversion",
            "Purchase conversion count","Purchase sales by conversion",
            "Cart conversion count","Cart sales by conversion",
            "Impression","Click","Cost","Average of Rank"
        ]:
            if col in df_combined.columns:
                df_combined[col] = pd.to_numeric(df_combined[col], errors="coerce").fillna(0)

        df_combined["Adgroup Type"] = df_combined["Adgroup Type"].astype(str).map(self.Adgroup_Type_map).fillna("기타")
        df_combined["Match Type"] = ""
        df_combined = df_combined.drop(columns=["Campaign ID","AD Group ID","Sum of Ad Rank"], errors="ignore")

        df_combined = df_combined.rename(columns={
            "Date": "일별",
            "Adgroup Type": "캠페인유형",
            "Campaign Name": "캠페인",
            "AD Group Name": "광고그룹",
            "Search keyword": "검색어",
            "Match Type": "검색유형",
            "Impression": "노출수",
            "Click": "클릭수",
            "Cost": "총비용(VAT포함,원)",
            "Total conversion count": "총 전환수",
            "Total sales by conversion": "총 전환매출액(원)",
            "Purchase conversion count": "구매완료 전환수",
            "Purchase sales by conversion": "구매완료 전환매출액(원)",
            "Cart conversion count": "장바구니 전환수",
            "Cart sales by conversion": "장바구니 전환매출액(원)",
            "Average of Rank": "평균 노출 순위"
        })

        df_criterion_keyword = pd.DataFrame(columns=df_combined.columns)

        sd = pd.to_datetime(start_date, format="%Y%m%d")
        month_label = f"{sd.month}월"
        week_of_month = ((sd.day - 1) // 7) + 1
        해당주차 = f"{week_of_month}주차"

        df_combined["일별"] = pd.to_datetime(df_combined["일별"], format="%Y%m%d", errors="coerce").dt.normalize()
        df_combined = df_combined.sort_values("일별")

        if not df_criterion_keyword.empty:
            df_combined = df_criterion_keyword.sort_values("일별")

        df_combined.insert(0, "해당월", month_label)
        df_combined.insert(1, "해당주차", 해당주차)
        df_combined.insert(2, "   ", "")
        df_combined.insert(3, "    ", "")
        df_combined.insert(4, "     ", "")

        # 3) PowerLink keywords
        final_columns2 = [
            "일별","캠페인유형","캠페인","광고그룹","검색어",
            "노출수","클릭수","총비용(VAT포함,원)",
            "총 전환수","총 전환매출액(원)",
            "구매완료 전환수","구매완료 전환매출액(원)",
            "평균 노출 순위",
            "장바구니 전환수","장바구니 전환매출액(원)",
            "검색유형"
        ]
        df_final2 = pd.DataFrame(columns=["해당월","해당주차","   ","    ","     "] + final_columns2)

        try:
            uri_kw = "/ncc/keywords"
            keywords_all = []
            for agid in df_ag["AD Group ID"].dropna().unique().tolist():
                r_kw = requests.get(
                    BASE_URL + uri_kw,
                    params={"nccAdgroupId": agid},
                    headers=Signature.get_header("GET", uri_kw, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID)
                )
                if r_kw.status_code == 200:
                    keywords_all.extend(r_kw.json())

            if not keywords_all:
                raise RuntimeError("파워링크 키워드가 없습니다(운영 X).")

            df_kw = (
                pd.DataFrame(keywords_all)[["nccKeywordId", "keyword", "nccAdgroupId"]]
                .rename(columns={
                    "nccKeywordId": "AD keyword ID",
                    "keyword": "Keyword",
                    "nccAdgroupId": "AD Group ID"
                })
            )

            df_kw_map = (
                df_kw.merge(df_ag[["AD Group ID","AD Group Name","Campaign ID","Adgroup Type"]], on="AD Group ID", how="left")
                    .merge(df_campaign[["Campaign ID","Campaign Name"]], on="Campaign ID", how="left")
                    [["Adgroup Type","Campaign ID","Campaign Name","AD Group ID","AD Group Name","AD keyword ID","Keyword"]]
            )

            ad_cols = [
                "Date","Customer_ID","Campaign ID","AD Group ID","AD keyword ID","AD ID","Business Channel ID",
                "Media Code","PC Mobile Type","Impression","Click","Cost","Sum of Ad Rank","View Count"
            ]
            ad_frames = []
            for d in self.daterange_yyyymmdd(start_date, end_date):
                ad_frames.append(self._safe_read_report_statdt("AD", d, ad_cols))
            df_ad = pd.concat(ad_frames, ignore_index=True) if ad_frames else self._empty_df(ad_cols)

            if df_ad.empty or df_ad["Campaign ID"].isna().all():
                raise RuntimeError("파워링크 AD 리포트가 없습니다(운영 X).")

            df_ad["Cost"] = pd.to_numeric(df_ad["Cost"], errors="coerce").fillna(0) 
            df_ad["Impression"] = pd.to_numeric(df_ad["Impression"], errors="coerce").fillna(0)
            df_ad["Click"] = pd.to_numeric(df_ad["Click"], errors="coerce").fillna(0)
            df_ad["Sum of Ad Rank"] = pd.to_numeric(df_ad["Sum of Ad Rank"], errors="coerce").fillna(0)

            df_ad = df_ad.merge(
                df_kw_map.drop_duplicates(subset=["Campaign ID","AD Group ID","AD keyword ID"]),
                on=["Campaign ID","AD Group ID","AD keyword ID"],
                how="left"
            )
            for c in ["Campaign Name","AD Group Name","Adgroup Type"]:
                df_ad[c] = df_ad[c].fillna("UNKNOWN")

            df_ad_g = (
                df_ad.groupby(
                    ["Date","Campaign Name","Campaign ID","AD Group ID","AD Group Name","Adgroup Type","AD keyword ID","Keyword"],
                    as_index=False
                )
                .agg({"Impression":"sum","Click":"sum","Cost":"sum","Sum of Ad Rank":"sum"})
            )
            df_ad_g["Average of Rank"] = np.where(
                df_ad_g["Impression"] > 0,
                df_ad_g["Sum of Ad Rank"] / df_ad_g["Impression"],
                np.nan
            ).round(1)

            conv_cols2 = [
                "Date","CUSTOMER ID","Campaign ID","AD Group ID","AD keyword ID","AD ID","Business Channel ID",
                "Hours","Region code","Media code","PC Mobile Type",
                "Conversion Method","Conversion Type","Conversion count","Sales by conversion"
            ]
            df_conv = self._safe_read_report_range("AD_CONVERSION_DETAIL", start_date, end_date, conv_cols2)
            if df_conv.empty:
                conv_frames = []
                for d in self.daterange_yyyymmdd(start_date, end_date):
                    conv_frames.append(self._safe_read_report_statdt("AD_CONVERSION_DETAIL", d, conv_cols2))
                df_conv = pd.concat(conv_frames, ignore_index=True) if conv_frames else self._empty_df(conv_cols2)

            df_conv = df_conv.merge(
                df_kw_map.drop_duplicates(subset=["Campaign ID","AD Group ID","AD keyword ID"]),
                on=["Campaign ID","AD Group ID","AD keyword ID"],
                how="left"
            )
            for c in ["Campaign Name","AD Group Name","Adgroup Type"]:
                df_conv[c] = df_conv[c].fillna("UNKNOWN")

            df_conv_g = self._group_conversion_breakdown(
                df_conv,
                ["Date","Campaign Name","Campaign ID","AD Group ID","AD Group Name","Adgroup Type","AD keyword ID","Keyword"]
            )

            mkeys = ["Date","Campaign Name","Campaign ID","AD Group ID","AD Group Name","Adgroup Type","AD keyword ID","Keyword"]
            for k in mkeys:
                df_conv_g[k] = df_conv_g[k].astype(str).str.strip()
                df_ad_g[k]   = df_ad_g[k].astype(str).str.strip()

            df_pl = pd.merge(df_conv_g, df_ad_g, on=mkeys, how="outer")

            for col in [
                "Total conversion count","Total sales by conversion",
                "Purchase conversion count","Purchase sales by conversion",
                "Cart conversion count","Cart sales by conversion",
                "Impression","Click","Cost","Average of Rank"
            ]:
                if col in df_pl.columns:
                    df_pl[col] = pd.to_numeric(df_pl[col], errors="coerce").fillna(0)

            df_pl["Adgroup Type"] = df_pl["Adgroup Type"].astype(str).map(self.Adgroup_Type_map).fillna("기타")
            df_pl["Search Keyword Type"] = "일치"

            df_pl = df_pl.rename(columns={
                "Date": "일별",
                "Adgroup Type": "캠페인유형",
                "Campaign Name": "캠페인",
                "AD Group Name": "광고그룹",
                "Keyword": "검색어",
                "Search Keyword Type": "검색유형",
                "Impression": "노출수",
                "Click": "클릭수",
                "Cost": "총비용(VAT포함,원)",
                "Total conversion count": "총 전환수",
                "Total sales by conversion": "총 전환매출액(원)",
                "Purchase conversion count": "구매완료 전환수",
                "Purchase sales by conversion": "구매완료 전환매출액(원)",
                "Cart conversion count": "장바구니 전환수",
                "Cart sales by conversion": "장바구니 전환매출액(원)",
                "Average of Rank": "평균 노출 순위"
            })

            df_pl["일별"] = pd.to_datetime(df_pl["일별"], format="%Y%m%d", errors="coerce").dt.normalize()
            df_pl = df_pl.sort_values("일별")

            exp_cols = [
                "Date","Customer ID","Campaign ID","AD Group ID","Search Keyword","Media code","PC Mobile Type",
                "Search Keyword Type","Impression","Click","Cost","View count"
            ]
            exp_frames = []
            for d in self.daterange_yyyymmdd(start_date, end_date):
                exp_frames.append(self._safe_read_report_statdt("EXPKEYWORD", d, exp_cols))
            df_exp = pd.concat(exp_frames, ignore_index=True) if exp_frames else self._empty_df(exp_cols)

            df_exp_final = pd.DataFrame(columns=final_columns2)
            if not df_exp.empty and not df_exp["Campaign ID"].isna().all():
                df_exp["Cost"] = pd.to_numeric(df_exp["Cost"], errors="coerce").fillna(0)
                df_exp["Impression"] = pd.to_numeric(df_exp["Impression"], errors="coerce").fillna(0)
                df_exp["Click"] = pd.to_numeric(df_exp["Click"], errors="coerce").fillna(0)
                df_exp["Search Keyword Type"] = df_exp["Search Keyword Type"].apply(
                    lambda x: self._normalize_search_type(x, default="확장")
                )
                df_exp = df_exp.merge(
                    df_cam.drop_duplicates(subset=["Campaign ID","AD Group ID"]),
                    on=["Campaign ID","AD Group ID"],
                    how="left"
                )
                for c in ["Campaign Name","AD Group Name","Adgroup Type"]:
                    df_exp[c] = df_exp[c].fillna("UNKNOWN")

                df_exp_g = (
                    df_exp.groupby(
                        ["Date","Campaign Name","Campaign ID","AD Group ID","AD Group Name","Adgroup Type","Search Keyword","Search Keyword Type"],
                        as_index=False
                    )
                    .agg({"Impression":"sum","Click":"sum","Cost":"sum"})
                )
                df_exp_g["Total conversion count"] = 0
                df_exp_g["Total sales by conversion"] = 0
                df_exp_g["Purchase conversion count"] = 0
                df_exp_g["Purchase sales by conversion"] = 0
                df_exp_g["Cart conversion count"] = 0
                df_exp_g["Cart sales by conversion"] = 0
                df_exp_g["Average of Rank"] = 0
                df_exp_g["Adgroup Type"] = df_exp_g["Adgroup Type"].astype(str).map(self.Adgroup_Type_map).fillna("기타")
                df_exp_g = df_exp_g.rename(columns={
                    "Date": "일별",
                    "Adgroup Type": "캠페인유형",
                    "Campaign Name": "캠페인",
                    "AD Group Name": "광고그룹",
                    "Search Keyword": "검색어",
                    "Search Keyword Type": "검색유형",
                    "Impression": "노출수",
                    "Click": "클릭수",
                    "Cost": "총비용(VAT포함,원)",
                    "Total conversion count": "총 전환수",
                    "Total sales by conversion": "총 전환매출액(원)",
                    "Purchase conversion count": "구매완료 전환수",
                    "Purchase sales by conversion": "구매완료 전환매출액(원)",
                    "Cart conversion count": "장바구니 전환수",
                    "Cart sales by conversion": "장바구니 전환매출액(원)",
                    "Average of Rank": "평균 노출 순위"
                })
                df_exp_g["일별"] = pd.to_datetime(df_exp_g["일별"], format="%Y%m%d", errors="coerce").dt.normalize()
                df_exp_final = df_exp_g[final_columns2].copy()

            df_final2 = pd.concat([df_pl[final_columns2].copy(), df_exp_final], ignore_index=True)
            df_final2.insert(0, "해당월", month_label)
            df_final2.insert(1, "해당주차", 해당주차)
            df_final2.insert(2, "   ", "")
            df_final2.insert(3, "    ", "")
            df_final2.insert(4, "     ", "")

        except Exception as e:
            self.log(f"파워링크 리포트 없음/미운영 -> 빈 DF 처리: {e}")

        # 4) Concat & Aggregate
        df_all = pd.concat([df_final2, df_combined], ignore_index=True)

        group_keys = ["캠페인유형", "캠페인", "광고그룹", "검색어", "검색유형"]
        sum_cols = [
            "노출수","클릭수","총비용(VAT포함,원)",
            "총 전환수","총 전환매출액(원)",
            "구매완료 전환수","구매완료 전환매출액(원)",
            "장바구니 전환수","장바구니 전환매출액(원)"
        ]

        for k in group_keys:
            if k in df_all.columns:
                df_all[k] = df_all[k].astype(str).str.strip()

        for c in sum_cols:
            if c in df_all.columns:
                df_all[c] = pd.to_numeric(df_all[c], errors="coerce").fillna(0)

        if "평균 노출 순위" in df_all.columns:
            df_all["평균 노출 순위"] = pd.to_numeric(df_all["평균 노출 순위"], errors="coerce").fillna(0)
            df_all["SumRank"] = df_all["평균 노출 순위"] * df_all["노출수"]
        else:
            df_all["SumRank"] = 0

        df_sum_base = df_all.copy()
        df_sum_base = df_sum_base[
            df_sum_base["검색어"].notna() &
            (df_sum_base["검색어"] != "") &
            (df_sum_base["검색어"] != "nan")
        ]

        df_total = (
            df_sum_base.groupby(group_keys, as_index=False)
            .agg({
                "노출수":"sum",
                "클릭수":"sum",
                "총비용(VAT포함,원)":"sum",
                "총 전환수":"sum",
                "총 전환매출액(원)":"sum",
                "구매완료 전환수":"sum",
                "구매완료 전환매출액(원)":"sum",
                "장바구니 전환수":"sum",
                "장바구니 전환매출액(원)":"sum",
                "SumRank":"sum"
            })
        )

        df_total["평균 노출 순위"] = np.where(
            df_total["노출수"] > 0,
            df_total["SumRank"] / df_total["노출수"],
            np.nan
        ).round(1)

        df_total = df_total.drop(columns=["SumRank"], errors="ignore")
        df_total = df_total.sort_values("총비용(VAT포함,원)", ascending=False).reset_index(drop=True)
        df_total = df_total[[
            "캠페인유형","캠페인","광고그룹","검색어",
            "노출수","클릭수","총비용(VAT포함,원)",
            "총 전환수","총 전환매출액(원)",
            "구매완료 전환수","구매완료 전환매출액(원)",
            "평균 노출 순위",
            "장바구니 전환수","장바구니 전환매출액(원)",
            "검색유형"
        ]]
        
        df_total.insert(0, "해당월", month_label)
        df_total.insert(1, "해당주차", 해당주차)
        df_total.insert(2, "   ", "")
        df_total.insert(3, "    ", "")
        df_total.insert(4, "     ", "")

        self._append_df_to_excel(df_total, excel, sheet_name)
        self.log("검색어 보고서 완료")

    def run_all_report_core(self, start_date: str, end_date: str, excel: str, daily_sheet: str, ad_sheet: str):
        self.log("일별+소재별 통합 보고서생성 시작!")
        BASE_URL = 'https://api.searchad.naver.com'

        # 1) Campaigns Mapping
        uri = "/ncc/campaigns"
        resp = requests.get(BASE_URL + uri, headers=Signature.get_header("GET", uri, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID))
        resp.raise_for_status()
        campaigns = resp.json()
        df_campaign = pd.DataFrame(campaigns)[["nccCampaignId", "name"]].rename(
            columns={"nccCampaignId": "Campaign ID", "name": "Campaign Name"}
        )

        # 2) Adgroups Mapping
        uri_ag = "/ncc/adgroups"
        adgroups_all = []
        for cid in df_campaign["Campaign ID"].tolist():
            r_ag = requests.get(
                BASE_URL + uri_ag,
                params={"nccCampaignId": cid},
                headers=Signature.get_header("GET", uri_ag, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID)
            )
            if r_ag.status_code == 200:
                adgroups_all.extend(r_ag.json())

        df_ag = pd.DataFrame(adgroups_all)[["nccAdgroupId", "name", "nccCampaignId", "adgroupType"]].rename(
            columns={
                "nccAdgroupId": "AD Group ID",
                "name": "AD Group Name",
                "nccCampaignId": "Campaign ID",
                "adgroupType": "Adgroup Type"
            }
        )

        # 3) Ads Mapping
        uri_ad = "/ncc/ads"
        ads_all = []
        for agid in df_ag["AD Group ID"].dropna().unique().tolist():
            r_ad = requests.get(
                BASE_URL + uri_ad,
                params={"nccAdgroupId": agid},
                headers=Signature.get_header("GET", uri_ad, self.API_KEY, self.SECRET_KEY, self.CUSTOMER_ID)
            )
            if r_ad.status_code == 200:
                ads_all.extend(r_ad.json())

        df_ad = pd.DataFrame(ads_all)[["nccAdId", "nccAdgroupId"]].rename(
            columns={"nccAdId": "AD ID", "nccAdgroupId": "AD Group ID"}
        )

        df_cam = (
            df_ag.merge(df_campaign[["Campaign ID", "Campaign Name"]], on="Campaign ID", how="left")
                .merge(df_ad, on="AD Group ID", how="left")
                [["Campaign Name","Campaign ID","AD Group Name","AD Group ID","AD ID","Adgroup Type"]]
        )

        # 4) AD report downloads
        column_names2 = [
            "Date", "Customer_ID", "Campaign ID","AD Group ID","AD keyword ID","AD ID","Business Channel ID",
            "Media Code","PC Mobile Type","Impression","Click","Cost","Sum of Ad Rank","View Count"
        ]

        frames = []
        for d in self.daterange_yyyymmdd(start_date, end_date):
            ad_text = self.create_and_download_report_statdt("AD", d)
            df_day = pd.read_csv(StringIO(ad_text), sep="\t", header=None, names=column_names2)
            frames.append(df_day)

        df_report2 = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=column_names2)

        df_report2["Cost"] = pd.to_numeric(df_report2["Cost"], errors="coerce").fillna(0) 
        df_report2["Impression"] = pd.to_numeric(df_report2["Impression"], errors="coerce").fillna(0)
        df_report2["Click"] = pd.to_numeric(df_report2["Click"], errors="coerce").fillna(0)
        df_report2['Average of Rank'] = df_report2['Sum of Ad Rank'] / df_report2['Impression']
        df_report2 = df_report2.round({'Average of Rank' : 1})

        df_report2 = df_report2.merge(
            df_cam.drop_duplicates(subset=["Campaign ID","AD Group ID","AD ID"]),
            on=["Campaign ID","AD Group ID","AD ID"],
            how="left"
        )

        for c in ["Campaign Name","AD Group Name","Adgroup Type"]:
            df_report2[c] = df_report2[c].fillna("UNKNOWN")

        df_report_group2 = df_report2.groupby(
                ["Date","Campaign Name","Campaign ID","AD Group ID","AD Group Name","Adgroup Type","AD ID","Media Code","PC Mobile Type"],
                as_index=False
            ) .agg({
                "Impression": "sum",
                "Click": "sum",
                "Cost": "sum",
                "Sum of Ad Rank": "sum",
                "View Count": "sum",
                "Average of Rank": "mean",
            })

        # 5) Conversion reports
        column_names = [
            "Date","CUSTOMER ID","Campaign ID","AD Group ID","AD keyword ID","AD ID","Business Channel ID",
            "Hours","Region code","Media code","PC Mobile Type",
            "Conversion Method","Conversion Type","Conversion count","Sales by conversion"
        ]

        df_report = pd.DataFrame(columns=column_names)
        try:
            conv_text = self.create_and_download_report_range("AD_CONVERSION_DETAIL", start_date, end_date)
            df_report = pd.read_csv(StringIO(conv_text), sep="\t", header=None, names=column_names)
        except Exception as e:
            self.log(f"[WARN] range 전환리포트 실패({start_date}~{end_date}) → statDt로 fallback. err={e}")

            frames = []
            for d in self.daterange_yyyymmdd(start_date, end_date):
                try:
                    conv_text = self.create_and_download_report_statdt("AD_CONVERSION_DETAIL", d)
                    df_day = pd.read_csv(StringIO(conv_text), sep="\t", header=None, names=column_names)
                    frames.append(df_day)
                    self.log(f'전환리포트(statDt) 성공 : {d}')
                except Exception as e2:
                    self.log(f"[SKIP] 전환리포트(statDt) 실패: {d}, err={e2}")
                    continue
            if frames:
                df_report = pd.concat(frames, ignore_index=True)

        df_report = df_report.merge(
            df_cam.drop_duplicates(subset=["Campaign ID","AD Group ID","AD ID"]),
            on=["Campaign ID","AD Group ID","AD ID"],
            how="left"
        )

        for c in ["Campaign Name","AD Group Name","Adgroup Type"]:
            df_report[c] = df_report[c].fillna("UNKNOWN")

        df_report_group = self._group_conversion_breakdown(
            df_report,
            ["Date","Campaign Name","Campaign ID","AD Group ID","AD Group Name","Adgroup Type","AD ID","Media code","PC Mobile Type"]
        )

        df_report_group = df_report_group.rename(columns={"Media code":"Media Code"})

        # 6) Merge
        merge_keys = ["Date","Campaign Name","Campaign ID","AD Group ID","AD Group Name","Adgroup Type","AD ID","Media Code","PC Mobile Type"]

        df_combined = pd.merge(df_report_group, df_report_group2, on=merge_keys, how="outer")

        for c in ["Total conversion count","Total sales by conversion","Purchase conversion count",
                  "Purchase sales by conversion","Cart conversion count","Cart sales by conversion"]:
            df_combined[c] = pd.to_numeric(df_combined[c], errors="coerce").fillna(0)
        df_combined["Impression"] = pd.to_numeric(df_combined["Impression"], errors="coerce").fillna(0)
        df_combined["Click"] = pd.to_numeric(df_combined["Click"], errors="coerce").fillna(0)
        df_combined["Cost"] = pd.to_numeric(df_combined["Cost"], errors="coerce").fillna(0)
        df_combined["Average of Rank"] = pd.to_numeric(df_combined["Average of Rank"], errors="coerce").fillna(0)
        df_combined["Media Name"] = pd.to_numeric(df_combined["Media Code"], errors="coerce").map(self.media_code_map).fillna("기타 매체")
        df_combined["Adgroup Type"] = df_combined["Adgroup Type"].astype(str).map(self.Adgroup_Type_map).fillna("기타")

        df_combined = df_combined.drop(columns=["Campaign ID","AD Group ID","Media Code"], errors="ignore")

        df_combined = df_combined.rename(columns={
            "Date": "일별",
            "Adgroup Type": "캠페인유형",
            "Campaign Name": "캠페인",
            "AD Group Name": "광고그룹",
            "AD ID": "소재",
            "PC Mobile Type": "PC/Mo",
            "Media Name": "매체이름",
            "Impression": "노출수",
            "Click": "클릭수",
            "Cost": "총비용(VAT포함,원)",
            "Total conversion count": "총 전환수",
            "Total sales by conversion": "총 전환매출액(원)",
            "Purchase conversion count": "구매완료 전환수",
            "Purchase sales by conversion": "구매완료 전환매출액(원)",
            "Cart conversion count": "장바구니 전환수",
            "Cart sales by conversion": "장바구니 전환매출액(원)",
            "Average of Rank": "평균 노출 순위"
        })
        sd = pd.to_datetime(start_date, format="%Y%m%d")

        month_label = f"{sd.month}월"
        week_of_month = ((sd.day - 1) // 7) + 1
        해당주차 = f"{week_of_month}주차"
                
        final_columns = [
            "일별","캠페인유형","캠페인","광고그룹","소재","PC/Mo",
            "노출수","클릭수","총비용(VAT포함,원)",
            "총 전환수","총 전환매출액(원)",
            "구매완료 전환수","구매완료 전환매출액(원)",
            "평균 노출 순위",
            "장바구니 전환수","장바구니 전환매출액(원)"
        ]

        final_columns = [c for c in final_columns if c in df_combined.columns]
        df_final = df_combined[final_columns].copy()

        df_final = (
            df_combined
            .groupby([
                "일별",
                "캠페인유형",
                "캠페인",
                "광고그룹",
                "소재",
                "PC/Mo"
            ], as_index=False)
            .agg({
                "노출수": "sum",
                "클릭수": "sum",
                "총비용(VAT포함,원)": "sum",
                "총 전환수": "sum",
                "총 전환매출액(원)": "sum",
                "구매완료 전환수": "sum",
                "구매완료 전환매출액(원)": "sum",
                "장바구니 전환수": "sum",
                "장바구니 전환매출액(원)": "sum",
                "평균 노출 순위": lambda x: np.average(
                    x,
                    weights=df_combined.loc[x.index, "노출수"]
                ) if df_combined.loc[x.index, "노출수"].sum() > 0 else np.nan
            })
        )
        if "평균 노출 순위" in df_final.columns:
            df_final["평균 노출 순위"] = (
                pd.to_numeric(df_final["평균 노출 순위"], errors="coerce")
                .round(1)
            )
        df_final = df_final[final_columns]
        df_final = df_final.sort_values("일별")
        df_final["일별"] = pd.to_datetime(df_final["일별"], format="%Y%m%d").dt.normalize()

        df_final.insert(0, "해당월", month_label)
        df_final.insert(1, "해당주차", 해당주차)
        df_final.insert(2, "   ", "")
        df_final.insert(3, "    ", "")
        df_final.insert(4, "     ", "")

        # Write daily report sheet
        self._append_df_to_excel(df_final, excel, daily_sheet)
        self.log("[OK] 일별 통합 데이터 시트 완료!")

        # Write ad/material report sheet
        value_cols = [
            "노출수", "클릭수", "총비용(VAT포함,원)",
            "총 전환수","총 전환매출액(원)",
            "구매완료 전환수","구매완료 전환매출액(원)",
            "장바구니 전환수","장바구니 전환매출액(원)"
        ]
        for c in value_cols:
            if c in df_combined.columns:
                df_combined[c] = pd.to_numeric(df_combined[c], errors="coerce").fillna(0)

        group_cols = ["캠페인유형", "캠페인", "광고그룹", "소재", "매체이름"]
        agg_map = {
            "노출수": "sum",
            "클릭수": "sum",
            "총비용(VAT포함,원)": "sum",
            "총 전환수": "sum",
            "총 전환매출액(원)": "sum",
            "구매완료 전환수": "sum",
            "구매완료 전환매출액(원)": "sum",
            "장바구니 전환수": "sum",
            "장바구니 전환매출액(원)": "sum",
        }
        if "평균 노출 순위" in df_combined.columns:
            agg_map["평균 노출 순위"] = "mean"

        df_ad_period = (
            df_combined
            .drop(columns=["일별", "PC/Mo"], errors="ignore")
            .groupby(group_cols, as_index=False)
            .agg(agg_map)
        )

        if "평균 노출 순위" in df_ad_period.columns:
            df_ad_period["평균 노출 순위"] = df_ad_period["평균 노출 순위"].round(1)
        df_ad_period = df_ad_period[[
            "캠페인유형","캠페인","광고그룹","소재","매체이름",
            "노출수","클릭수","총비용(VAT포함,원)",
            "총 전환수","총 전환매출액(원)",
            "구매완료 전환수","구매완료 전환매출액(원)",
            "평균 노출 순위",
            "장바구니 전환수","장바구니 전환매출액(원)"
        ]]

        df_ad_period.insert(0, "해당월", month_label)
        df_ad_period.insert(1, "해당주차", 해당주차)
        df_ad_period.insert(2, "   ", "")
        df_ad_period.insert(3, "    ", "")
        df_ad_period.insert(4, "     ", "")

        self._append_df_to_excel(df_ad_period, excel, ad_sheet)
        self.log("[OK] 소재별 통합 데이터 시트 완료!")


# ==========================================
# 5. Streamlit Application View & Controller
# ==========================================

# Main Header Design
st.markdown(
    """
    <div class="main-header-card">
        <h1 class="main-title">Naver Search AD Weekly Report</h1>
        <p class="subtitle">네이버 검색광고 API 기반 일별/소재별/검색어 보고서 자동화 시스템</p>
    </div>
    """,
    unsafe_allow_html=True
)

# Sidebar - API Credentials configuration
with st.sidebar:
    st.markdown("### 🔑 API 설정 정보")
    cust_id = st.text_input("CUSTOMER ID", value=st.session_state["customer_id"], key="customer_id", on_change=persist_state, placeholder="1596292")
    api_key = st.text_input("API KEY (Access License)", value=st.session_state["api_key"], key="api_key", on_change=persist_state, placeholder="Access License")
    sec_key = st.text_input("SECRET KEY", value=st.session_state["secret_key"], type="password", key="secret_key", on_change=persist_state, placeholder="Secret Key")
    
    st.markdown("---")
    st.markdown("### 💾 설정 보존")
    st.caption("입력한 데이터와 날짜 정보는 로컬 환경에 자동으로 암호화 및 저장되어 다음 번 재실행 시 유지됩니다.")

# Main Screen Layout
col1, col2 = st.columns(2)

# --- 1. Daily Report Card ---
with col1:
    st.markdown(
        """
        <div class="report-card">
            <div class="card-title">1. 쇼핑검색 - 일별 보고서</div>
        </div>
        """,
        unsafe_allow_html=True
    )
    with st.container():
        subcol1, subcol2 = st.columns(2)
        r1_s = subcol1.text_input("시작일 YYYYMMDD", value=st.session_state["r1_start"], key="r1_start", on_change=persist_state)
        r1_e = subcol2.text_input("종료일 YYYYMMDD", value=st.session_state["r1_end"], key="r1_end", on_change=persist_state)
        
        r1_excel = st.text_input("엑셀 파일 경로 (.xlsx)", value=st.session_state["r1_excel"], key="r1_excel", on_change=persist_state)
        r1_sheet = st.text_input("일별 시트 이름", value=st.session_state["r1_sheet"], key="r1_sheet", on_change=persist_state)
        
        r1_btn = st.button("일별 보고서 생성", key="r1_btn_run", use_container_width=True)

# --- 2. Material/AD Report Card ---
with col2:
    st.markdown(
        """
        <div class="report-card">
            <div class="card-title">2. 쇼핑검색 - 소재별 보고서</div>
        </div>
        """,
        unsafe_allow_html=True
    )
    with st.container():
        subcol1, subcol2 = st.columns(2)
        r2_s = subcol1.text_input("시작일 YYYYMMDD", value=st.session_state["r2_start"], key="r2_start", on_change=persist_state)
        r2_e = subcol2.text_input("종료일 YYYYMMDD", value=st.session_state["r2_end"], key="r2_end", on_change=persist_state)
        
        r2_excel = st.text_input("엑셀 파일 경로 (.xlsx)", value=st.session_state["r2_excel"], key="r2_excel", on_change=persist_state)
        r2_sheet = st.text_input("소재별 시트 이름", value=st.session_state["r2_sheet"], key="r2_sheet", on_change=persist_state)
        
        r2_btn = st.button("소재별 보고서 생성", key="r2_btn_run", use_container_width=True)

col3, col4 = st.columns(2)

# --- 3. Keyword Report Card ---
with col3:
    st.markdown(
        """
        <div class="report-card">
            <div class="card-title">3. 쇼핑검색 - 검색어 보고서</div>
        </div>
        """,
        unsafe_allow_html=True
    )
    with st.container():
        subcol1, subcol2 = st.columns(2)
        r3_s = subcol1.text_input("시작일 YYYYMMDD", value=st.session_state["r3_start"], key="r3_start", on_change=persist_state)
        r3_e = subcol2.text_input("종료일 YYYYMMDD", value=st.session_state["r3_end"], key="r3_end", on_change=persist_state)
        
        r3_excel = st.text_input("엑셀 파일 경로 (.xlsx)", value=st.session_state["r3_excel"], key="r3_excel", on_change=persist_state)
        r3_sheet = st.text_input("검색어 시트 이름", value=st.session_state["r3_sheet"], key="r3_sheet", on_change=persist_state)
        
        r3_btn = st.button("검색어 보고서 생성", key="r3_btn_run", use_container_width=True)

# --- 4. Master Combined Action Card ---
with col4:
    st.markdown(
        """
        <div class="report-card" style="border-color: rgba(245, 158, 11, 0.4); background: rgba(245, 158, 11, 0.02);">
            <div class="card-title" style="border-left-color: #f59e0b;">🌟 전체 마스터 보고서 일괄 실행</div>
        </div>
        """,
        unsafe_allow_html=True
    )
    with st.container():
        subcol1, subcol2 = st.columns(2)
        all_s = subcol1.text_input("시작일 YYYYMMDD", value=st.session_state["all_start"], key="all_start", on_change=persist_state)
        all_e = subcol2.text_input("종료일 YYYYMMDD", value=st.session_state["all_end"], key="all_end", on_change=persist_state)
        
        all_excel = st.text_input("엑셀 전체 보고서 저장 경로 (.xlsx)", value=st.session_state["all_excel"], key="all_excel", on_change=persist_state)
        
        st.markdown("<div style='height: 14px;'></div>", unsafe_allow_html=True)
        r4_btn = st.button("⚡ 전체 보고서 실행 버튼 (일별 + 소재별 + 검색어 일괄)", key="r4_btn_run", use_container_width=True)

# Console Output Area Setup
st.markdown("### 💻 실시간 실행 콘솔 로그")
console_placeholder = st.empty()

# Custom console initial status rendering
console_placeholder.markdown(
    """
    <div class="console-box">
        <div class="console-header">
            <span class="dot red"></span>
            <span class="dot yellow"></span>
            <span class="dot green"></span>
            <span class="console-title">Console Logs</span>
        </div>
        <div class="console-content">
            [SYS] Naver Search AD Web App Initialized. Ready to generate reports.
        </div>
    </div>
    """,
    unsafe_allow_html=True
)

# Date validation helper
def validate_date(s: str) -> bool:
    try:
        datetime.strptime(s, "%Y%m%d")
        return True
    except ValueError:
        return False

# ==========================================
# 6. Streamlit Event Loop and Routing Actions
# ==========================================
if r1_btn:
    if not api_key or not sec_key or not cust_id:
        st.error("⚠️ API 입력 정보(Customer ID, Access License, Secret Key)를 먼저 완벽히 채워주세요.")
    elif not validate_date(r1_s) or not validate_date(r1_e):
        st.error("⚠️ 날짜 형식이 바르지 않습니다. YYYYMMDD 형태로 입력해 주세요.")
    elif r1_s > r1_e:
        st.error("⚠️ 시작일은 종료일보다 클 수 없습니다.")
    else:
        with capture_stdout(console_placeholder):
            try:
                runner = NaverReportRunner(api_key, sec_key, cust_id)
                st.info("🔄 일별 보고서를 수집하고 병합하는 중...")
                runner.run_daily_report(r1_s, r1_e, r1_sheet, r1_excel)
                st.success("🎉 일별 보고서 생성이 완료되었습니다!")
                
                # Render direct download link if file exists
                if os.path.exists(r1_excel):
                    with open(r1_excel, "rb") as f:
                        st.download_button(
                            label="📥 생성된 엑셀 파일 바로 다운로드",
                            data=f,
                            file_name=os.path.basename(r1_excel),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            except Exception as e:
                print(f"[FATAL ERROR] {e}")
                traceback.print_exc()
                st.error(f"❌ 보고서 생성 중 치명적인 에러가 발생했습니다: {e}")

if r2_btn:
    if not api_key or not sec_key or not cust_id:
        st.error("⚠️ API 입력 정보(Customer ID, Access License, Secret Key)를 먼저 완벽히 채워주세요.")
    elif not validate_date(r2_s) or not validate_date(r2_e):
        st.error("⚠️ 날짜 형식이 바르지 않습니다. YYYYMMDD 형태로 입력해 주세요.")
    elif r2_s > r2_e:
        st.error("⚠️ 시작일은 종료일보다 클 수 없습니다.")
    else:
        with capture_stdout(console_placeholder):
            try:
                runner = NaverReportRunner(api_key, sec_key, cust_id)
                st.info("🔄 소재별 보고서를 수집하고 병합하는 중...")
                runner.run_ad_report(r2_s, r2_e, r2_sheet, r2_excel)
                st.success("🎉 소재별 보고서 생성이 완료되었습니다!")
                
                if os.path.exists(r2_excel):
                    with open(r2_excel, "rb") as f:
                        st.download_button(
                            label="📥 생성된 엑셀 파일 바로 다운로드",
                            data=f,
                            file_name=os.path.basename(r2_excel),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            except Exception as e:
                print(f"[FATAL ERROR] {e}")
                traceback.print_exc()
                st.error(f"❌ 보고서 생성 중 치명적인 에러가 발생했습니다: {e}")

if r3_btn:
    if not api_key or not sec_key or not cust_id:
        st.error("⚠️ API 입력 정보(Customer ID, Access License, Secret Key)를 먼저 완벽히 채워주세요.")
    elif not validate_date(r3_s) or not validate_date(r3_e):
        st.error("⚠️ 날짜 형식이 바르지 않습니다. YYYYMMDD 형태로 입력해 주세요.")
    elif r3_s > r3_e:
        st.error("⚠️ 시작일은 종료일보다 클 수 없습니다.")
    else:
        with capture_stdout(console_placeholder):
            try:
                runner = NaverReportRunner(api_key, sec_key, cust_id)
                st.info("🔄 검색어 보고서를 수집하고 가중 순위를 계산하는 중...")
                runner.run_keyword_report(r3_s, r3_e, r3_sheet, r3_excel)
                st.success("🎉 검색어 보고서 생성이 완료되었습니다!")
                
                if os.path.exists(r3_excel):
                    with open(r3_excel, "rb") as f:
                        st.download_button(
                            label="📥 생성된 엑셀 파일 바로 다운로드",
                            data=f,
                            file_name=os.path.basename(r3_excel),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            except Exception as e:
                print(f"[FATAL ERROR] {e}")
                traceback.print_exc()
                st.error(f"❌ 보고서 생성 중 치명적인 에러가 발생했습니다: {e}")

if r4_btn:
    if not api_key or not sec_key or not cust_id:
        st.error("⚠️ API 입력 정보(Customer ID, Access License, Secret Key)를 먼저 완벽히 채워주세요.")
    elif not validate_date(all_s) or not validate_date(all_e):
        st.error("⚠️ 날짜 형식이 바르지 않습니다. YYYYMMDD 형태로 입력해 주세요.")
    elif all_s > all_e:
        st.error("⚠️ 시작일은 종료일보다 클 수 없습니다.")
    else:
        with capture_stdout(console_placeholder):
            try:
                runner = NaverReportRunner(api_key, sec_key, cust_id)
                
                # Step 1: Run Combined Daily + Material/AD Reports
                st.info("🔄 1. 일별 및 소재별 통합 보고서 실행 중...")
                runner.run_all_report_core(all_s, all_e, all_excel, r1_sheet, r2_sheet)
                
                # Step 2: Run Keyword Report
                st.info("🔄 2. 검색어 보고서 실행 중...")
                runner.run_keyword_report(all_s, all_e, r3_sheet, all_excel)
                
                st.success("🎉 전체 보고서 일괄 생성이 완벽히 완료되었습니다!")
                
                if os.path.exists(all_excel):
                    with open(all_excel, "rb") as f:
                        st.download_button(
                            label="📥 생성된 전체 통합 엑셀 파일 바로 다운로드",
                            data=f,
                            file_name=os.path.basename(all_excel),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            except Exception as e:
                print(f"[FATAL ERROR] {e}")
                traceback.print_exc()
                st.error(f"❌ 전체 보고서 일괄 실행 중 치명적인 에러가 발생했습니다: {e}")
